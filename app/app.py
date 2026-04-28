import os
import zipfile
from datetime import datetime, date, time as dt_time
from io import BytesIO
import calendar

import streamlit as st
import pandas as pd
import pymssql
import oracledb
import os

try:
    lib_dir = [d for d in os.listdir('/opt') if d.startswith('instantclient')][0]
    oracledb.init_oracle_client(lib_dir=f"/opt/{lib_dir}")
except Exception as e:
    print(f"Erro ao iniciar Oracle Client: {e}")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════════════════
def _cfg(section: str, key: str, default=None):
    """Lê de st.secrets (deploy) ou variável de ambiente (local)."""
    try:
        return st.secrets[section][key]
    except Exception:
        return os.getenv(key, default)

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "Logos", "Via Appia", "PNG", "Via Appia Positivo.png")

DB_HOST    = _cfg("oracle", "DB_HOST",    "oracluster")
DB_PORT    = int(_cfg("oracle", "DB_PORT", "1521"))
DB_SERVICE = _cfg("oracle", "DB_SERVICE", "srv_bi")
DB_USER    = _cfg("oracle", "DB_USER")
DB_PASS    = _cfg("oracle", "DB_PASSWORD")
SK_EMPRESA = int(_cfg("oracle", "SK_EMPRESA", "6"))

try:
    oracledb.init_oracle_client()
except Exception:
    pass

SQL_HOST = _cfg("sqlserver", "SQL_HOST", "192.168.40.22")
SQL_PORT = int(_cfg("sqlserver", "SQL_PORT", "1433"))
SQL_DB   = _cfg("sqlserver", "SQL_DB",   "AT3_db_2")
SQL_USER = _cfg("sqlserver", "SQL_USER", "sa")
SQL_PASS = _cfg("sqlserver", "SQL_PASS", "$4dmp4dr40@nasc")

CONCESSIONARIA = "Via Nascentes"

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

COR_HEADER_BG = "1B5D5D"
COR_HEADER_FG = "FFFFFF"
COR_ALT_ROW   = "D6EDED"

# ══════════════════════════════════════════════════════════════════════════════
# CSS DO COCKPIT
# ══════════════════════════════════════════════════════════════════════════════
CSS = """
<style>
:root {
    --verde:    #1b5d5d;
    --laranja:  #ff5400;
    --agua:     #3ca4a6;
    --fundo:    #f0f6f6;
    --card-bg:  #ffffff;
    --texto:    #1a1a1a;
}

.stApp { background-color: var(--fundo); }

h1, h2, h3 { color: var(--verde) !important; }

/* Barra de navegação superior */
.nav-bar {
    display: flex;
    gap: 8px;
    margin-bottom: 24px;
    flex-wrap: wrap;
}

/* Card de automação */
.card {
    background: var(--card-bg);
    border-radius: 12px;
    padding: 20px 22px;
    border-left: 5px solid var(--agua);
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    height: 100%;
    transition: box-shadow 0.2s;
}
.card:hover { box-shadow: 0 4px 16px rgba(27,93,93,0.15); }
.card-icon  { font-size: 2rem; margin-bottom: 8px; }
.card-title { font-size: 1rem; font-weight: 700; color: var(--verde); margin: 0 0 6px 0; }
.card-desc  { font-size: 0.82rem; color: #555; margin: 0; line-height: 1.4; }

/* Card de modelo */
.modelo-card {
    background: var(--card-bg);
    border-radius: 10px;
    padding: 14px 18px;
    border-left: 4px solid var(--laranja);
    box-shadow: 0 1px 5px rgba(0,0,0,0.06);
    margin-bottom: 8px;
}
.modelo-titulo { font-size: 0.88rem; font-weight: 600; color: var(--verde); }
.modelo-sub    { font-size: 0.76rem; color: #888; }

/* Breadcrumb */
.breadcrumb {
    font-size: 0.8rem;
    color: var(--agua);
    margin-bottom: 4px;
}

/* Botões primários */
.stButton > button[kind="primary"],
.stButton > button[data-testid="stBaseButton-primary"] {
    background-color: var(--laranja) !important;
    border: none !important;
    color: #fff !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
}
.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="stBaseButton-primary"]:hover {
    background-color: #e04b00 !important;
}

/* Botões de download */
.stDownloadButton > button {
    background-color: var(--agua) !important;
    border: none !important;
    color: #fff !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
}
.stDownloadButton > button:hover { background-color: #2e8b8d !important; }

/* Progress bar */
.stProgress > div > div > div > div { background-color: var(--laranja) !important; }

/* Texto geral */
.stApp, .stApp p, .stApp span, .stApp div,
.stMarkdown, .stMarkdown p, .stMarkdown span,
[data-testid="stText"], [data-testid="stCaptionContainer"] p {
    color: var(--texto) !important;
}

/* Métricas */
[data-testid="stMetricValue"] { color: var(--verde) !important; font-weight: 700 !important; }
[data-testid="stMetricLabel"] p { color: var(--texto) !important; font-weight: 600 !important; }
[data-testid="stMetricDelta"]   { color: #555 !important; }

/* Progress bar — texto do label */
[data-testid="stProgress"] p,
[data-testid="stProgressBar"] + p,
.stProgress p { color: var(--texto) !important; }

/* Alertas: info, warning, error, success */
[data-testid="stAlert"] p,
[data-testid="stAlert"] span,
[data-testid="stNotification"] p,
[data-testid="stNotification"] span,
div[class*="stInfo"] p,
div[class*="stWarning"] p,
div[class*="stError"] p,
div[class*="stSuccess"] p {
    color: var(--texto) !important;
}

/* Spinner */
[data-testid="stSpinner"] p { color: var(--texto) !important; }

/* Inputs */
.stSelectbox > div > div,
.stTextInput > div > div > input,
.stNumberInput > div > div > input {
    border-color: var(--agua) !important;
    border-radius: 6px !important;
}

/* Labels dos inputs */
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] label,
.stSelectbox label,
.stTextInput label,
.stNumberInput label {
    color: var(--texto) !important;
}

/* Badge de pasta */
.badge {
    display: inline-block;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.04em;
    padding: 2px 8px;
    border-radius: 20px;
    margin-bottom: 10px;
    text-transform: uppercase;
}
.badge-op   { background: #e6f4f1; color: #1b5d5d; }
.badge-sats { background: #fff0e8; color: #cc4400; }

/* Divider */
hr { border-color: var(--agua) !important; opacity: 0.35; }

/* Dataframe header */
.stDataFrame thead th { background-color: var(--verde) !important; color: #fff !important; }
</style>
"""

# ══════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════════════════
def safe(val) -> str:
    try:
        if pd.isna(val): return ""
    except Exception:
        pass
    return str(val).strip()

def safe_int(val) -> int:
    try:
        return int(float(str(val).strip()))
    except Exception:
        return 0

def safe_float(val) -> float:
    try:
        return float(str(val).strip())
    except Exception:
        return 0.0

def fmt_valor(v) -> str:
    try:
        return "{:.2f}".format(float(v)).replace(".", ",")
    except Exception:
        return "0,00"

def _norm(txt: str) -> str:
    txt = str(txt).strip().lower()
    for a, b in [
        ("á","a"),("à","a"),("â","a"),("ã","a"),
        ("é","e"),("ê","e"),("í","i"),("ó","o"),
        ("ô","o"),("õ","o"),("ú","u"),("ç","c"),
    ]:
        txt = txt.replace(a, b)
    return txt

def formatar_datahora(data_val, hora_val) -> str:
    try:
        if isinstance(data_val, datetime):   dt = data_val.date()
        elif isinstance(data_val, date):     dt = data_val
        else: dt = datetime.strptime(str(data_val).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return ""
    try:
        if isinstance(hora_val, dt_time):    hs = hora_val.strftime("%H:%M:%S")
        elif hora_val is None or (isinstance(hora_val, float) and pd.isna(hora_val)): hs = "00:00:00"
        else:
            hs = str(hora_val).strip()
            if len(hs) == 5: hs += ":00"
    except Exception:
        hs = "00:00:00"
    return "{}-{}".format(dt.strftime("%d/%m/%Y"), hs)

def fmt_km(km, mt) -> str:
    try:
        return "{}+{}".format(
            str(int(float(str(km).strip()))).zfill(3),
            str(int(float(str(mt).strip()))).zfill(3),
        )
    except Exception:
        return str(km).strip()

def get_oracle_connection():
    dsn = oracledb.makedsn(DB_HOST, DB_PORT, service_name=DB_SERVICE)
    return oracledb.connect(user=DB_USER, password=DB_PASS, dsn=dsn)

def get_sqlserver_connection():
    return pymssql.connect(
        server=SQL_HOST,
        port=SQL_PORT,
        database=SQL_DB,
        user=SQL_USER,
        password=SQL_PASS,
    )

def estilizar_excel(wb, ws, colunas, dados_df):
    hf   = Font(name="Arial", bold=True, color=COR_HEADER_FG, size=10)
    hb   = PatternFill("solid", start_color=COR_HEADER_BG)
    ha   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dfnt = Font(name="Arial", size=9)
    af   = PatternFill("solid", start_color=COR_ALT_ROW)
    ca   = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 35
    for ci, col in enumerate(colunas, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf; c.fill = hb; c.alignment = ha; c.border = brd
    for ri, row in enumerate(dados_df[colunas].itertuples(index=False), 2):
        fill = af if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfnt; c.alignment = ca; c.border = brd
            if fill: c.fill = fill
    ws.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════════════
# ATENDIMENTOS
# ══════════════════════════════════════════════════════════════════════════════
COLUNAS_ATENDIMENTOS = [
    "CnpjConcessionaria","NumOcorrencia","IdVeiculo","TipoVeiculo",
    "DataAcionamento","KmMetrosAcionamento","LatitudeAcionamento","LongitudeAcionamento",
    "DataChegadaAtendimento","DataSaidaAtendimento",
    "LatitudeAtendimento","LongitudeAtendimento",
    "Rodovia","KmMetrosAtendimento",
    "ViaUrbanaNaoSegregada","Pavimentacao","TipoServicoPrestado",
]

def map_tipo_veiculo_atend(viatura: str) -> str:
    v = str(viatura).strip().upper()
    if v.startswith("IT-"):                           return "Inspeção de tráfego"
    if v.startswith("GL-"):                           return "Guincho leve"
    if v.startswith("GPD-"):                          return "Guincho pesado"
    if v.startswith("B-"):                            return "Ambulância Tipo C"
    if v.startswith("P-"):                            return "Caminhão pipa"
    if v.startswith("OPE-") or v.startswith("PGF-"): return "Inspeção de tráfego"
    return ""

def map_tipo_servico(tipo_atend: str, viatura: str) -> str:
    ta = str(tipo_atend).strip().lower()
    v  = str(viatura).strip().upper()
    if ta == "cancelado":                                         return "Cancelado"
    if "nao localizado" in ta or "não localizado" in ta:         return "Não localizado"
    if v.startswith("IT-") or v.startswith("OPE-") or v.startswith("PGF-"):
        return "Inspeção (Veículo de inspeção)"
    if v.startswith("GL-"):  return "Mecânico (Guincho leve)"
    if v.startswith("GPD-"): return "Mecânico (Guincho pesado)"
    if v.startswith("B-"):   return "Atendimento pré-hospitalar (Ambulância)"
    if v.startswith("P-"):   return "Combate a incêndio (Caminhão pipa)"
    if "inspe" in ta:        return "Inspeção (Veículo de inspeção)"
    if "remo"  in ta:        return "Mecânico (Guincho leve)"
    return ""

QUERY_ATENDIMENTOS = """
    SELECT
        KOCORRENCIA, NUMOCORRENCIA, DATAOCORRENCIA, HORAOCORRENCIA,
        COD_RODOVIA, RODOVIA, KM, COD_SENT, SENTIDO,
        COD_TIPO_ATENDIMENTO, DSC_TIPO_ATENDIMENTO, TIPO_REMOCAO,
        RECURSO, TEMPOCHEGADA, TIPOOCORRENCIA, DESCROCORRENCIA,
        LATITUDE, LONGITUDE, TRECHO, ANO, MES, DIA
    FROM STAGE_AREA.STG_ATENDIMENTOS_KCOR
    WHERE SK_EMPRESA = :sk_empresa
      AND ANO        = :ano
      AND MES        = :mes
    ORDER BY DATAOCORRENCIA, HORAOCORRENCIA
"""

def carregar_atendimentos(mes, ano):
    try:
        conn = get_oracle_connection()
        df = pd.read_sql(QUERY_ATENDIMENTOS, conn, params={"sk_empresa": SK_EMPRESA, "ano": ano, "mes": mes})
        conn.close()
        return df
    except Exception as e:
        st.error("Erro Oracle: {}".format(e))
        return pd.DataFrame()

def transformar_atendimentos(df):
    if df.empty: return pd.DataFrame(columns=COLUNAS_ATENDIMENTOS)
    rows = []
    for _, r in df.iterrows():
        viatura    = safe(r.get("RECURSO"))
        tipo_atend = safe(r.get("DSC_TIPO_ATENDIMENTO"))
        rows.append({
            "CnpjConcessionaria":     CONCESSIONARIA,
            "NumOcorrencia":          safe(r.get("NUMOCORRENCIA")),
            "IdVeiculo":              viatura,
            "TipoVeiculo":            map_tipo_veiculo_atend(viatura),
            "DataAcionamento":        formatar_datahora(r.get("DATAOCORRENCIA"), r.get("HORAOCORRENCIA")),
            "KmMetrosAcionamento":    "",
            "LatitudeAcionamento":    "",
            "LongitudeAcionamento":   "",
            "DataChegadaAtendimento": formatar_datahora(r.get("DATAOCORRENCIA"), r.get("TEMPOCHEGADA")),
            "DataSaidaAtendimento":   "",
            "LatitudeAtendimento":    safe(r.get("LATITUDE")),
            "LongitudeAtendimento":   safe(r.get("LONGITUDE")),
            "Rodovia":                safe(r.get("RODOVIA")),
            "KmMetrosAtendimento":    safe(r.get("KM")),
            "ViaUrbanaNaoSegregada":  "",
            "Pavimentacao":           "",
            "TipoServicoPrestado":    map_tipo_servico(tipo_atend, viatura),
        })
    return pd.DataFrame(rows, columns=COLUNAS_ATENDIMENTOS)

def gerar_xlsx_atendimentos(df):
    wb = Workbook(); ws = wb.active; ws.title = "Atendimentos Serviços"
    estilizar_excel(wb, ws, COLUNAS_ATENDIMENTOS, df)
    widths = [22,16,12,24,24,22,20,20,26,24,20,20,12,22,24,14,34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def salvar_atendimentos(xlsx_bytes, mes, ano, base_dir):
    pasta = os.path.join(base_dir, "A-OPERACAO", "atendimentos_servicos",
                         str(ano), "{}_{}".format(str(mes).zfill(2), MESES_PT[mes]))
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, "{}_{}_Atendimentos_Servicos.xlsx".format(ano, str(mes).zfill(2)))
    with open(caminho, "wb") as f: f.write(xlsx_bytes)
    return caminho

# ══════════════════════════════════════════════════════════════════════════════
# SINISTROS
# ══════════════════════════════════════════════════════════════════════════════
COLUNAS_SINISTROS = [
    "CnpjConcessionaria","NumOcorrencia","IdEnvolvido","DataHora",
    "Municipio","Sre","Rodovia","KmMetros","Sentido",
    "Latitude","Longitude","TipoPista","TracadoVia",
    "CondicaoMetereologica","Iluminacao","ObstrucaoPista",
    "CausaPrincipal","CausaSinistro","OrdemTipoSinistro","TipoSinistro",
    "ClassificacaoSinistro","QuantidadeVeiculos",
    "IdVeiculo","TipoVeiculoEnvolvidoSinistro","MarcaModeloVeiculo",
    "AnoFabricacaoVeiculo","PlacaVeiculo","TipoCarga",
    "QuantidadeEnvolvidos","TipoEnvolvido","EstadoFisico",
    "IdadeEnvolvido","SexoEnvolvido",
    "Ileso","FeridoLeve","FeridoModerado","FeridoGrave","Morto",
]

SENTIDO_MAP = {
    "leste":"Crescente","norte":"Crescente","l":"Crescente","n":"Crescente",
    "oeste":"Decrescente","sul":"Decrescente","o":"Decrescente","s":"Decrescente",
}
TIPO_SINISTRO_MAP = {
    "1.0 colisao traseira":"Colisão traseira",
    "2.0 colisao frontal":"Colisão frontal",
    "3.0 colisao lateral":"Colisão lateral",
    "6.6 atrop. pedestre morador/trabalhador/estudante":"Atropelamento de pedestre",
    "7.0 atropelamento de animal":"Atropelamento de animal",
    "8.0 tombamento":"Tombamento",
    "9.0 capotamento":"Capotamento",
    "10.0 engavetamento":"Engavetamento",
    "5.0 saida de pista":"Saída de pista",
}
TRACADO_MAP = {"reta":"Reta","curva":"Curva"}
CONDICAO_TEMPO_MAP = {"bom":"Bom","chuva":"Chuva","nublado":"Nublado","nevoeiro":"Nevoeiro"}
CAUSA_MAP = {
    "velocidade incompativel":"Velocidade incompatível",
    "nao guardar distancia de seguranca":"Não guardar distância",
    "desobediencia a sinalizacao":"Desobediência à sinalização",
    "dormindo":"Dormindo",
    "ingestao de alcool":"Ingestão de álcool",
    "defeito mecanico em veiculo":"Defeito mecânico",
    "animais na pista":"Animais na pista",
    "falta de atencao":"Falta de atenção",
    "ultrapassagem indevida":"Ultrapassagem indevida",
}
CLASSIFICACAO_MAP = {
    "acidente sem vitima":"Sem vítima",
    "acidente com vitima":"Com vítima",
    "acidente com vitima fatal":"Fatal",
}

def map_tipo_veiculo_sinistro(desc):
    d = _norm(desc)
    if "automovel" in d or "automov" in d:          return "Automóvel"
    if "caminhonete" in d or "utilitario" in d:     return "Caminhonete/Utilitário"
    if "caminhao" in d:                             return "Caminhão"
    if "moto" in d:                                 return "Motocicleta"
    if "onibus" in d or "micro" in d:               return "Ônibus"
    if "bicicleta" in d:                            return "Bicicleta"
    return desc

QUERY_ACIDENTES = """
    SELECT
        A.KOCORRENCIA, A.NUMOCORRENCIA, A.DATAOCORRENCIA, A.HORAACIDENTE,
        A.DESCROCORRENCIA, A.TIPOACIDENTE, A.CAUSAPROVAVEL,
        A.TRACADOPISTA, A.CONDICAOMETEOROLOGICA,
        A.LATITUDE, A.LONGITUDE, A.SIGLARODOVIA, A.KM, A.MTS, A.TRECHO,
        MAX(AT.SENTIDO) AS SENTIDO
    FROM STAGE_AREA.STG_KCOR_ACIDENTES2 A
    LEFT JOIN STAGE_AREA.STG_ATENDIMENTOS_KCOR AT
        ON AT.KOCORRENCIA = A.KOCORRENCIA AND AT.SK_EMPRESA = A.SK_EMPRESA
    WHERE A.SK_EMPRESA = :sk_empresa
      AND EXTRACT(YEAR  FROM A.DATAOCORRENCIA) = :ano
      AND EXTRACT(MONTH FROM A.DATAOCORRENCIA) = :mes
    GROUP BY A.KOCORRENCIA, A.NUMOCORRENCIA, A.DATAOCORRENCIA, A.HORAACIDENTE,
             A.DESCROCORRENCIA, A.TIPOACIDENTE, A.CAUSAPROVAVEL,
             A.TRACADOPISTA, A.CONDICAOMETEOROLOGICA,
             A.LATITUDE, A.LONGITUDE, A.SIGLARODOVIA, A.KM, A.MTS, A.TRECHO
    ORDER BY A.DATAOCORRENCIA, A.HORAACIDENTE
"""

QUERY_VEICULOS = """
    SELECT V.KOCORRENCIA, V.COD_VEICULO_ATEND, V.DESC_VEICULO_ATEN, V.MARCAMODELO
    FROM STAGE_AREA.STG_KCOR_VEICULOS V
    WHERE V.SK_EMPRESA = :sk_empresa
      AND EXTRACT(YEAR  FROM V.DATAOCORRENCIA) = :ano
      AND EXTRACT(MONTH FROM V.DATAOCORRENCIA) = :mes
    ORDER BY V.KOCORRENCIA, V.COD_VEICULO_ATEND
"""

def carregar_sinistros(mes, ano):
    try:
        conn   = get_oracle_connection()
        params = {"sk_empresa": SK_EMPRESA, "ano": ano, "mes": mes}
        df_acid = pd.read_sql(QUERY_ACIDENTES, conn, params=params)
        df_veic = pd.read_sql(QUERY_VEICULOS,  conn, params=params)
        conn.close()
        return df_acid, df_veic
    except Exception as e:
        st.error("Erro Oracle: {}".format(e))
        return pd.DataFrame(), pd.DataFrame()

def transformar_sinistros(df_acid, df_veic):
    if df_acid.empty: return pd.DataFrame(columns=COLUNAS_SINISTROS)
    qtd_veic = (
        df_veic.groupby("KOCORRENCIA")["COD_VEICULO_ATEND"]
        .count().reset_index().rename(columns={"COD_VEICULO_ATEND": "QTD_VEIC"})
    ) if not df_veic.empty else pd.DataFrame(columns=["KOCORRENCIA","QTD_VEIC"])
    rows = []
    for _, r in df_acid.iterrows():
        koco          = r["KOCORRENCIA"]
        data_hora     = formatar_datahora(r.get("DATAOCORRENCIA"), r.get("HORAACIDENTE"))
        sentido       = SENTIDO_MAP.get(_norm(safe(r.get("SENTIDO"))), safe(r.get("SENTIDO")))
        tracado       = TRACADO_MAP.get(_norm(safe(r.get("TRACADOPISTA"))), safe(r.get("TRACADOPISTA")))
        cond_met      = CONDICAO_TEMPO_MAP.get(_norm(safe(r.get("CONDICAOMETEOROLOGICA"))), safe(r.get("CONDICAOMETEOROLOGICA")))
        causa         = CAUSA_MAP.get(_norm(safe(r.get("CAUSAPROVAVEL"))), safe(r.get("CAUSAPROVAVEL")))
        tipo_sinistro = TIPO_SINISTRO_MAP.get(_norm(safe(r.get("TIPOACIDENTE"))), safe(r.get("TIPOACIDENTE")))
        classificacao = CLASSIFICACAO_MAP.get(_norm(safe(r.get("DESCROCORRENCIA"))), safe(r.get("DESCROCORRENCIA")))
        km_metros     = fmt_km(r.get("KM",""), r.get("MTS",0))
        descr_norm    = _norm(safe(r.get("DESCROCORRENCIA")))
        qtd_row = qtd_veic[qtd_veic["KOCORRENCIA"]==koco]
        qtd_v   = int(qtd_row["QTD_VEIC"].values[0]) if not qtd_row.empty else ""
        veics_oco = (
            df_veic[df_veic["KOCORRENCIA"]==koco].sort_values("COD_VEICULO_ATEND").reset_index(drop=True)
        ) if not df_veic.empty else pd.DataFrame()
        if veics_oco.empty:
            veics_oco = pd.DataFrame([{"COD_VEICULO_ATEND":1,"DESC_VEICULO_ATEN":"","MARCAMODELO":""}])
        for i, vei in veics_oco.iterrows():
            ileso=ferido_leve=ferido_mod=ferido_grave=morto=0
            if i==0:
                if "sem vitima" in descr_norm:   ileso=1
                elif "fatal" in descr_norm:      morto=1
                elif "com vitima" in descr_norm: ferido_leve=1
            rows.append({
                "CnpjConcessionaria":CONCESSIONARIA,"NumOcorrencia":safe(r.get("NUMOCORRENCIA")),
                "IdEnvolvido":safe(vei.get("COD_VEICULO_ATEND")),"DataHora":data_hora,
                "Municipio":"","Sre":safe(r.get("TRECHO")),"Rodovia":safe(r.get("SIGLARODOVIA")),
                "KmMetros":km_metros,"Sentido":sentido,"Latitude":safe(r.get("LATITUDE")),
                "Longitude":safe(r.get("LONGITUDE")),"TipoPista":"","TracadoVia":tracado,
                "CondicaoMetereologica":cond_met,"Iluminacao":"","ObstrucaoPista":"",
                "CausaPrincipal":"Sim" if i==0 else "Não","CausaSinistro":causa if i==0 else "",
                "OrdemTipoSinistro":str(i+1),"TipoSinistro":tipo_sinistro,
                "ClassificacaoSinistro":classificacao,"QuantidadeVeiculos":qtd_v,
                "IdVeiculo":safe(vei.get("COD_VEICULO_ATEND")),
                "TipoVeiculoEnvolvidoSinistro":map_tipo_veiculo_sinistro(safe(vei.get("DESC_VEICULO_ATEN"))),
                "MarcaModeloVeiculo":safe(vei.get("MARCAMODELO")),
                "AnoFabricacaoVeiculo":"","PlacaVeiculo":"","TipoCarga":"",
                "QuantidadeEnvolvidos":"","TipoEnvolvido":"","EstadoFisico":"",
                "IdadeEnvolvido":"","SexoEnvolvido":"",
                "Ileso":ileso,"FeridoLeve":ferido_leve,"FeridoModerado":ferido_mod,
                "FeridoGrave":ferido_grave,"Morto":morto,
            })
    return pd.DataFrame(rows, columns=COLUNAS_SINISTROS)

def gerar_xlsx_sinistros(df):
    wb = Workbook(); ws = wb.active; ws.title = "Sinistros Trânsito"
    estilizar_excel(wb, ws, COLUNAS_SINISTROS, df)
    widths = [22,16,12,24,18,12,12,14,14,16,16,14,16,22,14,18,16,26,20,26,24,20,14,28,24,20,14,14,20,16,16,14,14,8,10,12,10,8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def salvar_sinistros(xlsx_bytes, mes, ano, base_dir):
    pasta = os.path.join(base_dir, "A-OPERACAO", "sinistros_transito",
                         str(ano), "{}_{}".format(str(mes).zfill(2), MESES_PT[mes]))
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, "{}_{}_Sinistros_Transito.xlsx".format(ano, str(mes).zfill(2)))
    with open(caminho, "wb") as f: f.write(xlsx_bytes)
    return caminho

# ══════════════════════════════════════════════════════════════════════════════
# TRÁFEGO
# ══════════════════════════════════════════════════════════════════════════════
COLUNAS_TRAFEGO = [
    "CnpjConcessionaria","IdPassagem","CodigoPracaPedagio","CodigoCabine",
    "DataHoraPassagem","Faixa","Sentido","TipoVeiculo",
    "QuantidadeEixosVeiculo","CategoriaVeiculo","Rodagem",
    "EixoSuspenso","QuantidadeEixosSuspensos",
    "Isento","TipoCobrancaEfetuada","Placa",
    "ValorDevido","ValorArrecadado","IdTag",
]

def map_tipo_veiculo_trafego(passeio, comercial, moto):
    p=safe_int(passeio); c=safe_int(comercial); m=safe_int(moto)
    if m>0: return "Moto"
    if c>0: return "Comercial"
    if p>0: return "Passeio"
    return ""

def map_tipo_cobranca(subforma):
    s = str(subforma).strip().upper()
    if "TAG" in s:              return "Automática TAG"
    if "OCR" in s or "PLACA" in s: return "Automática OCR/PLACA"
    if "MAN" in s:              return "Manual"
    return safe(subforma)

def map_sentido_trafego(dsc_sentido):
    s = str(dsc_sentido).strip().upper()
    if "CRESC" in s or "LESTE" in s or "NORTE" in s: return "Crescente"
    if "DECRESC" in s or "OESTE" in s or "SUL" in s: return "Decrescente"
    return safe(dsc_sentido)

QUERY_TRAFEGO_DIA = """
    SELECT
        f.COD_ID_TRANSACTION, f.SK_VALIDADOR, f.NUM_HORA_PASSAGEM,
        f.DAC_POSDER, f.DAC_POSIZQ, f.QTD_PASSEIO, f.QTD_COMERCIAL, f.QTD_MOTO,
        f.DAC_ADICION, f.DAC_CLASSE, f.COD_CAT_VEIC, f.COD_CLASSE,
        f.NUM_OTICO_NUM_EIXOS_SUSP, f.DSC_SUBFORMA_PGTO,
        f.SK_FORMA_PAGAMENTO, f.VAL_VALOR_ARREC,
        d.DAT_REF_FORMATADA, p.COD_PRACA, p.DSC_PRACA, p.DSC_SENTIDO, p.COD_VIA
    FROM DW.FAT_ARRECADACAO_TRAFEGO f
    LEFT JOIN DW.DIM_DAT_PERIODO d ON f.SK_DAT_PERIODO = d.SK_DAT_PERIODO
    LEFT JOIN DW.DIM_EMPRESA_PRACA_VIA p ON f.SK_EMPRESA_PRACA_VIA = p.SK_EMPRESA_PRACA_VIA
    WHERE p.DSC_EMPRESA = 'NASCENTES DAS GERAIS'
      AND d.DAT_REF_FORMATADA >= TO_DATE(:dt_dia,      'YYYY-MM-DD')
      AND d.DAT_REF_FORMATADA <  TO_DATE(:dt_dia_prox, 'YYYY-MM-DD')
    ORDER BY f.NUM_HORA_PASSAGEM
"""

QUERY_EVASOES_DIA = """
    SELECT
        ev.IDTRANSACTION, ev.PLACA,
        f.SK_VALIDADOR, f.NUM_HORA_PASSAGEM,
        f.DAC_POSDER, f.DAC_POSIZQ, f.QTD_PASSEIO, f.QTD_COMERCIAL, f.QTD_MOTO,
        f.DAC_ADICION, f.DAC_CLASSE, f.COD_CAT_VEIC, f.COD_CLASSE,
        f.NUM_OTICO_NUM_EIXOS_SUSP, f.DSC_SUBFORMA_PGTO,
        f.SK_FORMA_PAGAMENTO, f.COD_ID_TRANSACTION,
        d.DAT_REF_FORMATADA, p.COD_PRACA, p.DSC_PRACA, p.DSC_SENTIDO
    FROM DW.FAT_EVASOES ev
    LEFT JOIN DW.FAT_ARRECADACAO_TRAFEGO f
        ON TO_CHAR(f.COD_ID_TRANSACTION) = TO_CHAR(ev.IDTRANSACTION)
    LEFT JOIN DW.DIM_DAT_PERIODO d ON f.SK_DAT_PERIODO = d.SK_DAT_PERIODO
    LEFT JOIN DW.DIM_EMPRESA_PRACA_VIA p ON f.SK_EMPRESA_PRACA_VIA = p.SK_EMPRESA_PRACA_VIA
    WHERE p.DSC_EMPRESA = 'NASCENTES DAS GERAIS'
      AND d.DAT_REF_FORMATADA >= TO_DATE(:dt_dia,      'YYYY-MM-DD')
      AND d.DAT_REF_FORMATADA <  TO_DATE(:dt_dia_prox, 'YYYY-MM-DD')
    ORDER BY f.NUM_HORA_PASSAGEM
"""

def transformar_trafego(df, tarifas):
    rows = []
    for _, r in df.iterrows():
        val_arrec  = safe_float(r.get("VAL_VALOR_ARREC"))
        eixos_susp = safe_int(r.get("NUM_OTICO_NUM_EIXOS_SUSP"))
        eixos_tot  = safe_int(r.get("DAC_ADICION")) + safe_int(r.get("DAC_CLASSE"))
        rows.append({
            "CnpjConcessionaria":       CONCESSIONARIA,
            "IdPassagem":               safe(r.get("COD_ID_TRANSACTION")),
            "CodigoPracaPedagio":       safe(r.get("COD_PRACA")),
            "CodigoCabine":             safe(r.get("SK_VALIDADOR")),
            "DataHoraPassagem":         formatar_datahora(r.get("DAT_REF_FORMATADA"), r.get("NUM_HORA_PASSAGEM")),
            "Faixa":                    safe(r.get("DAC_POSDER")),
            "Sentido":                  map_sentido_trafego(safe(r.get("DSC_SENTIDO"))),
            "TipoVeiculo":              map_tipo_veiculo_trafego(r.get("QTD_PASSEIO"), r.get("QTD_COMERCIAL"), r.get("QTD_MOTO")),
            "QuantidadeEixosVeiculo":   eixos_tot if eixos_tot>0 else "",
            "CategoriaVeiculo":         safe(r.get("COD_CAT_VEIC")),
            "Rodagem":                  "Simples",
            "EixoSuspenso":             "Sim" if eixos_susp>0 else "Não",
            "QuantidadeEixosSuspensos": eixos_susp if eixos_susp>0 else "",
            "Isento":                   "Sim" if val_arrec==0 else "Não",
            "TipoCobrancaEfetuada":     map_tipo_cobranca(safe(r.get("DSC_SUBFORMA_PGTO"))),
            "Placa":                    "",
            "ValorDevido":              fmt_valor(val_arrec),
            "ValorArrecadado":          fmt_valor(val_arrec),
            "IdTag":                    safe(r.get("SK_FORMA_PAGAMENTO")),
        })
    return pd.DataFrame(rows, columns=COLUNAS_TRAFEGO)

def transformar_evasoes(df, tarifas):
    rows = []
    for _, r in df.iterrows():
        cod_praca  = safe(r.get("COD_PRACA"))
        tarifa_ref = tarifas.get(cod_praca, 0.0)
        eixos_susp = safe_int(r.get("NUM_OTICO_NUM_EIXOS_SUSP"))
        eixos_tot  = safe_int(r.get("DAC_ADICION")) + safe_int(r.get("DAC_CLASSE"))
        rows.append({
            "CnpjConcessionaria":       CONCESSIONARIA,
            "IdPassagem":               safe(r.get("IDTRANSACTION")),
            "CodigoPracaPedagio":       cod_praca,
            "CodigoCabine":             safe(r.get("SK_VALIDADOR")),
            "DataHoraPassagem":         formatar_datahora(r.get("DAT_REF_FORMATADA"), r.get("NUM_HORA_PASSAGEM")),
            "Faixa":                    safe(r.get("DAC_POSDER")),
            "Sentido":                  map_sentido_trafego(safe(r.get("DSC_SENTIDO"))),
            "TipoVeiculo":              map_tipo_veiculo_trafego(r.get("QTD_PASSEIO"), r.get("QTD_COMERCIAL"), r.get("QTD_MOTO")),
            "QuantidadeEixosVeiculo":   eixos_tot if eixos_tot>0 else "",
            "CategoriaVeiculo":         safe(r.get("COD_CAT_VEIC")),
            "Rodagem":                  "Simples",
            "EixoSuspenso":             "Sim" if eixos_susp>0 else "Não",
            "QuantidadeEixosSuspensos": eixos_susp if eixos_susp>0 else "",
            "Isento":                   "Não",
            "TipoCobrancaEfetuada":     "Evasão",
            "Placa":                    safe(r.get("PLACA")),
            "ValorDevido":              fmt_valor(tarifa_ref),
            "ValorArrecadado":          fmt_valor(0.0),
            "IdTag":                    "",
        })
    return pd.DataFrame(rows, columns=COLUNAS_TRAFEGO)

def caminho_csv_trafego(pasta_base, mes, ano):
    pasta = os.path.join(pasta_base, "A-OPERACAO", "trafego_praca_pedagio",
                         str(ano), "{}_{}".format(str(mes).zfill(2), MESES_PT[mes]))
    os.makedirs(pasta, exist_ok=True)
    return os.path.join(pasta, "{}_{}_Trafego_Praca_Pedagio.csv".format(ano, str(mes).zfill(2)))

def gerar_trafego_por_dia(mes, ano, pasta_base):
    from datetime import timedelta
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dias       = [date(ano, mes, d) for d in range(1, ultimo_dia + 1)]
    total_dias = len(dias)
    caminho    = caminho_csv_trafego(pasta_base, mes, ano)
    tarifas    = {}
    total_pass = 0; total_ev = 0; total_perd = 0.0; primeiro = True

    if os.path.exists(caminho): os.remove(caminho)

    conn     = get_oracle_connection()
    barra    = st.progress(0, text="Iniciando...")
    status_tx= st.empty()
    col1, col2, col3 = st.columns(3)
    m_pass = col1.empty(); m_ev = col2.empty(); m_perd = col3.empty()

    try:
        for idx, dia in enumerate(dias):
            dia_str      = dia.strftime("%Y-%m-%d")
            dia_prox_str = (dia + timedelta(days=1)).strftime("%Y-%m-%d")
            params       = {"dt_dia": dia_str, "dt_dia_prox": dia_prox_str}
            status_tx.info("Processando {}/{} ({}/{})...".format(
                dia.strftime("%d/%m/%Y"), ano, idx+1, total_dias))

            df_t = pd.read_sql(QUERY_TRAFEGO_DIA, conn, params=params)
            df_pago = df_t[df_t["VAL_VALOR_ARREC"] > 0]
            for praca, grupo in df_pago.groupby("COD_PRACA"):
                chave = str(praca)
                if chave not in tarifas:
                    tarifas[chave] = round(grupo["VAL_VALOR_ARREC"].median(), 2)

            df_e = pd.read_sql(QUERY_EVASOES_DIA, conn, params=params)
            frames = []
            if not df_t.empty:
                frames.append(transformar_trafego(df_t, tarifas))
                total_pass += len(df_t)
            if not df_e.empty:
                frames.append(transformar_evasoes(df_e, tarifas))
                total_perd += sum(tarifas.get(safe(r.get("COD_PRACA")), 0.0) for _, r in df_e.iterrows())
                total_ev   += len(df_e)

            if frames:
                df_dia = pd.concat(frames, ignore_index=True)
                df_dia.to_csv(caminho, mode="w" if primeiro else "a",
                              header=primeiro, index=False, sep=";", encoding="utf-8-sig")
                primeiro = False

            pct = int(((idx+1)/total_dias)*100)
            barra.progress(pct, text="{}% — Dia {}/{} concluído".format(pct, idx+1, total_dias))
            m_pass.metric("Passagens",     "{:,}".format(total_pass))
            m_ev.metric("Evasões",         "{:,}".format(total_ev))
            m_perd.metric("Valor Perdido", "R$ {}".format(fmt_valor(total_perd)))
    except Exception as e:
        st.error("Erro no processamento: {}".format(e))
    finally:
        conn.close()

    return caminho, total_pass, total_ev, total_perd

# ══════════════════════════════════════════════════════════════════════════════
# SATS
# ══════════════════════════════════════════════════════════════════════════════
TABELAS_SATS = [
    "AT3_1_Amostra","AT3_1_Ocorrencia",
    "AT3_2_Amostra","AT3_2_Ocorrencia",
    "AT3_3_Amostra","AT3_3_Ocorrencia",
    "AT3_4_Amostra","AT3_4_Ocorrencia",
    "AT3_5_Amostra","AT3_5_Ocorrencia",
    "AT3_6_Amostra","AT3_6_Ocorrencia",
    "AT3_7_Amostra","AT3_7_Ocorrencia",
    "AT3_8_Amostra","AT3_8_Ocorrencia",
    "AT3_9_Amostra","AT3_9_Ocorrencia",
    "AT3_10_Amostra","AT3_10_Ocorrencia",
    "AT3_11_Amostra","AT3_11_Ocorrencia",
    "AT3_12_Amostra","AT3_12_Ocorrencia",
    "AT3_13_Amostra","AT3_13_Ocorrencia",
    "AT3_14_Amostra","AT3_14_Ocorrencia",
    "AT3_15_Amostra","AT3_15_Ocorrencia",
    "AT3_16_Amostra","AT3_16_Ocorrencia",
    "AT3_17_Amostra","AT3_17_Ocorrencia",
    "AT3_18_Amostra","AT3_18_Ocorrencia",
    "AT3_19_Amostra","AT3_19_Ocorrencia",
    "AT3_20_Amostra","AT3_20_Ocorrencia",
    "AT3_37_Amostra","AT3_37_Ocorrencia",
    "AT3_99_Amostra","AT3_99_Ocorrencia",
    "AT3_231_Amostra","AT3_231_Ocorrencia",
]

def _sats_df_to_xlsx(df):
    buf = BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()

def _sats_build_zip(resultados, mes, ano):
    prefixo = "{}_{}_{:04d}".format(str(mes).zfill(2), MESES_PT[mes], ano)
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nome, xlsx_bytes in resultados.items():
            pasta = "AMOSTRAS" if nome.endswith("_Amostra") else "OCORRENCIAS"
            zf.writestr("{}/{}_{}.xlsx".format(pasta, nome, prefixo), xlsx_bytes)
    return buf.getvalue()

def gerar_sats(mes, ano):
    ultimo_dia  = calendar.monthrange(ano, mes)[1]
    data_inicio = date(ano, mes, 1)
    data_fim    = date(ano, mes, ultimo_dia)
    total       = len(TABELAS_SATS)
    barra       = st.progress(0, text="Iniciando...")
    status_tx   = st.empty()
    col1, col2, col3 = st.columns(3)
    m_am = col1.empty(); m_oc = col2.empty(); m_li = col3.empty()
    qtd_am=0; qtd_oc=0; total_linhas=0; resultados={}

    try:
        conn = get_sqlserver_connection()
        for i, nome_tabela in enumerate(TABELAS_SATS):
            status_tx.info("Processando **{}** ({}/{})...".format(nome_tabela, i+1, total))
            query = """
                SELECT * FROM AT3_db_1.dbo.{tabela}
                WHERE 1=1
                  AND Instante >= '{dt_ini}'
                  AND Instante <= '{dt_fim}'
                ORDER BY Instante DESC
            """.format(tabela=nome_tabela, dt_ini=data_inicio, dt_fim=data_fim)
            cursor = conn.cursor()
            cursor.execute(query)
            cols = [d[0] for d in cursor.description]
            df   = pd.DataFrame(cursor.fetchall(), columns=cols)
            total_linhas += len(df)
            if nome_tabela.endswith("_Amostra"): qtd_am += 1
            else:                                qtd_oc += 1
            resultados[nome_tabela] = _sats_df_to_xlsx(df)
            pct = int(((i+1)/total)*100)
            barra.progress(pct, text="{}% — {}/{} tabelas".format(pct, i+1, total))
            m_am.metric("Amostras",      qtd_am)
            m_oc.metric("Ocorrências",   qtd_oc)
            m_li.metric("Linhas totais", "{:,}".format(total_linhas))
        conn.close()
        status_tx.empty()
        st.success("Concluído! {} amostras | {} ocorrências | {:,} linhas totais.".format(
            qtd_am, qtd_oc, total_linhas))
    except pymssql.Error as e:
        st.error("Erro SQL Server: {}".format(e))
    except Exception as e:
        st.error("Erro: {}".format(e))

    return resultados

def render_downloads_sats(resultados, mes, ano):
    if not resultados: return
    prefixo   = "{}_{}_{:04d}".format(str(mes).zfill(2), MESES_PT[mes], ano)
    mime_xlsx = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    st.download_button(
        label="Baixar TODOS os relatórios SATS (ZIP)",
        data=_sats_build_zip(resultados, mes, ano),
        file_name="CSATS_{}.zip".format(prefixo),
        mime="application/zip",
        use_container_width=True,
        type="primary",
    )
    st.divider()

    amostras    = {k:v for k,v in resultados.items() if k.endswith("_Amostra")}
    ocorrencias = {k:v for k,v in resultados.items() if not k.endswith("_Amostra")}
    col_a, col_o = st.columns(2)
    with col_a:
        st.markdown("**Amostras**")
        for nome, xlsx_bytes in amostras.items():
            st.download_button(
                label=nome, data=xlsx_bytes,
                file_name="{}_{}.xlsx".format(nome, prefixo),
                mime=mime_xlsx, use_container_width=True,
                key="dl_sats_{}".format(nome),
            )
    with col_o:
        st.markdown("**Ocorrências**")
        for nome, xlsx_bytes in ocorrencias.items():
            st.download_button(
                label=nome, data=xlsx_bytes,
                file_name="{}_{}.xlsx".format(nome, prefixo),
                mime=mime_xlsx, use_container_width=True,
                key="dl_sats_{}".format(nome),
            )

# ══════════════════════════════════════════════════════════════════════════════
# MODELOS DE PLANILHA
# ══════════════════════════════════════════════════════════════════════════════
MODELOS = [
    ("Atendimentos e Serviços",          "A-OPERAÇÃO",       "atendimentos.xlsx"),
    ("Sinistros de Trânsito",            "A-OPERAÇÃO",       "sinistros_transito.xlsx"),
    ("Tráfego Praça de Pedágio",         "A-OPERAÇÃO",       "trafego_praca_pedagio.xlsx"),
    ("Localização de Ativos",            "A-OPERAÇÃO",       "localizacao_ativos_manual.xlsx"),
    ("Apreensão de Animais",             "A-OPERAÇÃO",       "apreensao_animais.xlsx"),
    ("Tráfego SAT / Contador Veicular",  "C-SATs",           "trafego_sat_contador_veicular.xlsx"),
    ("Disponibilidade de Equipamentos",  "D-DISPONIBILIDADE","disponibilidade_equipamentos_manual.xlsx"),
    ("Disponibilidade Operacional",      "D-DISPONIBILIDADE","disponibilidade_operacional_pesagem_manual.xlsx"),
    ("Pesagem Veicular",                 "E-PESAGEMVEICULAR","pesagem_veicular.xlsx"),
]

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ══════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS DE UI
# ══════════════════════════════════════════════════════════════════════════════
def nav(label, destino):
    if st.button(label, use_container_width=True, key="nav_{}".format(destino)):
        st.session_state["pagina"] = destino
        st.rerun()

def voltar():
    if st.button("← Voltar ao Cockpit", key="voltar_cockpit"):
        st.session_state["pagina"] = "cockpit"
        st.rerun()

def cabecalho(titulo, icone, subtitulo=""):
    col_logo, col_txt = st.columns([1, 4])
    with col_logo:
        if os.path.exists(LOGO_PATH):
            st.image(LOGO_PATH, width=120)
    with col_txt:
        st.markdown(
            '<p class="breadcrumb">ARTEMIG — Cockpit de Relatórios</p>',
            unsafe_allow_html=True,
        )
        st.markdown("# {} {}".format(icone, titulo))
        if subtitulo:
            st.caption(subtitulo)
    st.divider()

def seletor_mes_ano(sufixo):
    col1, col2 = st.columns(2)
    with col1:
        mes = st.selectbox(
            "Mês", options=list(MESES_PT.keys()),
            format_func=lambda m: MESES_PT[m],
            index=datetime.now().month - 1,
            key="mes_{}".format(sufixo),
        )
    with col2:
        ano = st.number_input(
            "Ano", min_value=2020, max_value=datetime.now().year + 1,
            value=datetime.now().year, step=1,
            key="ano_{}".format(sufixo),
        )
    return mes, int(ano)

# ══════════════════════════════════════════════════════════════════════════════
# PÁGINAS
# ══════════════════════════════════════════════════════════════════════════════

def pagina_cockpit():
    if os.path.exists(LOGO_PATH):
        col = st.columns([1, 2, 1])
        with col[1]:
            st.image(LOGO_PATH, use_container_width=True)
    st.markdown(
        "<h1 style='text-align:center;margin-top:8px;'>Cockpit de Relatórios — ARTEMIG</h1>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p style='text-align:center;color:#3ca4a6;margin-top:-8px;'>Via Nascentes · Painel central de emissão de relatórios e download de modelos</p>",
        unsafe_allow_html=True,
    )
    st.divider()

    # ── Automações ───────────────────────────────────────────────────────────
    st.markdown("## Automações")
    st.caption("Clique em uma automação para configurar e gerar o relatório.")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div class="card">
            <span class="badge badge-op">📁 A-OPERAÇÃO</span>
            <div class="card-icon">🚑</div>
            <p class="card-title">Atendimentos e Serviços</p>
            <p class="card-desc">Extrai atendimentos operacionais do Oracle (STAGE_AREA) e gera planilha Excel padronizada.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Gerar Atendimentos", key="ir_atend", use_container_width=True, type="primary"):
            st.session_state["pagina"] = "atendimentos"
            st.rerun()

    with col2:
        st.markdown("""
        <div class="card">
            <span class="badge badge-op">📁 A-OPERAÇÃO</span>
            <div class="card-icon">🚗</div>
            <p class="card-title">Sinistros de Trânsito</p>
            <p class="card-desc">Cruza tabelas de acidentes e veículos no Oracle e gera relatório de ocorrências por envolvido.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Gerar Sinistros", key="ir_sinis", use_container_width=True, type="primary"):
            st.session_state["pagina"] = "sinistros"
            st.rerun()

    st.write("")

    col3, col4 = st.columns(2)
    with col3:
        st.markdown("""
        <div class="card">
            <span class="badge badge-op">📁 A-OPERAÇÃO</span>
            <div class="card-icon">🛣️</div>
            <p class="card-title">Tráfego Praça de Pedágio</p>
            <p class="card-desc">Processa passagens e evasões dia a dia do DW Oracle e salva CSV incremental.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Gerar Tráfego", key="ir_traf", use_container_width=True, type="primary"):
            st.session_state["pagina"] = "trafego"
            st.rerun()

    with col4:
        st.markdown("""
        <div class="card">
            <span class="badge badge-sats">📁 C-SATS</span>
            <div class="card-icon">📊</div>
            <p class="card-title">C-SATS — Extração AT3</p>
            <p class="card-desc">Consulta 28 tabelas AT3 (Amostras e Ocorrências) no SQL Server e gera arquivos Excel por tipo.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Gerar SATS", key="ir_sats", use_container_width=True, type="primary"):
            st.session_state["pagina"] = "sats"
            st.rerun()

    st.divider()

    # ── Modelos de planilha ──────────────────────────────────────────────────
    st.markdown("## Modelos de Planilha")
    st.caption("Baixe a planilha modelo de cada tipo de relatório para preenchimento manual.")

    colunas = st.columns(3)
    for idx, (nome, pasta, arquivo) in enumerate(MODELOS):
        caminho = os.path.join(BASE_DIR, pasta, arquivo)
        with colunas[idx % 3]:
            st.markdown(
                '<div class="modelo-card">'
                '<p class="modelo-titulo">{}</p>'
                '<p class="modelo-sub">{}</p>'
                '</div>'.format(nome, pasta),
                unsafe_allow_html=True,
            )
            if os.path.exists(caminho):
                with open(caminho, "rb") as f:
                    st.download_button(
                        label="Baixar modelo",
                        data=f.read(),
                        file_name="Modelo_{}.xlsx".format(nome.replace(" ","_").replace("/","-")),
                        mime=MIME_XLSX,
                        use_container_width=True,
                        key="modelo_{}".format(idx),
                    )
            else:
                st.caption("Arquivo não encontrado")


def pagina_atendimentos():
    cabecalho("Atendimentos e Serviços", "🚑",
              "Extrai dados de STAGE_AREA.STG_ATENDIMENTOS_KCOR e gera Excel padronizado.")
    voltar()
    st.write("")

    mes, ano = seletor_mes_ano("atend")
    st.info("Oracle · SK_EMPRESA={} · {}/{}".format(SK_EMPRESA, str(mes).zfill(2), ano))

    pasta_base = st.text_input(
        "Pasta de destino (opcional)",
        value=os.path.join(os.path.expanduser("~"), "Relatórios Nascentes"),
        key="pasta_atend",
    )

    st.write("")
    if st.button("Gerar Relatório de Atendimentos", type="primary", use_container_width=True, key="btn_atend"):
        with st.spinner("Consultando Oracle..."):
            df_raw = carregar_atendimentos(mes, ano)
        if df_raw.empty:
            st.warning("Nenhum registro encontrado.")
            return
        with st.spinner("Transformando dados..."):
            df_final = transformar_atendimentos(df_raw)
        st.success("{} registros gerados.".format(len(df_final)))
        st.dataframe(df_final, use_container_width=True)
        xlsx_bytes = gerar_xlsx_atendimentos(df_final)
        nome_arquivo = "{}_{}_Atendimentos_Servicos.xlsx".format(ano, str(mes).zfill(2))
        st.download_button(
            "Baixar Relatório Excel", data=xlsx_bytes,
            file_name=nome_arquivo, mime=MIME_XLSX,
            use_container_width=True,
        )
        if pasta_base:
            try:
                caminho = salvar_atendimentos(xlsx_bytes, mes, ano, pasta_base)
                st.success("Salvo em: {}".format(caminho))
            except Exception as e:
                st.warning("Não foi possível salvar localmente: {}".format(e))


def pagina_sinistros():
    cabecalho("Sinistros de Trânsito", "🚗",
              "Cruza STG_KCOR_ACIDENTES2 + STG_KCOR_VEICULOS e gera relatório por envolvido.")
    voltar()
    st.write("")

    mes, ano = seletor_mes_ano("sinis")
    st.info("Oracle · SK_EMPRESA={} · {}/{}".format(SK_EMPRESA, str(mes).zfill(2), ano))

    pasta_base = st.text_input(
        "Pasta de destino (opcional)",
        value=os.path.join(os.path.expanduser("~"), "Relatórios Nascentes"),
        key="pasta_sinis",
    )

    st.write("")
    if st.button("Gerar Relatório de Sinistros", type="primary", use_container_width=True, key="btn_sinis"):
        with st.spinner("Consultando Oracle..."):
            df_acid, df_veic = carregar_sinistros(mes, ano)
        if df_acid.empty:
            st.warning("Nenhum acidente encontrado.")
            return
        with st.spinner("Cruzando tabelas..."):
            df_final = transformar_sinistros(df_acid, df_veic)
        st.success("{} ocorrências → {} linhas geradas.".format(df_acid.shape[0], df_final.shape[0]))
        st.dataframe(df_final, use_container_width=True)
        xlsx_bytes = gerar_xlsx_sinistros(df_final)
        nome_arquivo = "{}_{}_Sinistros_Transito.xlsx".format(ano, str(mes).zfill(2))
        st.download_button(
            "Baixar Relatório Excel", data=xlsx_bytes,
            file_name=nome_arquivo, mime=MIME_XLSX,
            use_container_width=True,
        )
        if pasta_base:
            try:
                caminho = salvar_sinistros(xlsx_bytes, mes, ano, pasta_base)
                st.success("Salvo em: {}".format(caminho))
            except Exception as e:
                st.warning("Não foi possível salvar localmente: {}".format(e))


def pagina_trafego():
    cabecalho("Tráfego Praça de Pedágio", "🛣️",
              "Processa passagens e evasões dia a dia. CSV salvo incrementalmente.")
    voltar()
    st.write("")

    mes, ano = seletor_mes_ano("traf")
    n_dias = calendar.monthrange(ano, mes)[1]
    st.info("Oracle DW · Empresa: NASCENTES DAS GERAIS · {}/{} · {} dias".format(
        str(mes).zfill(2), ano, n_dias))

    pasta_base = st.text_input(
        "Pasta de destino (obrigatório para tráfego)",
        value=os.path.join(os.path.expanduser("~"), "Relatórios Nascentes"),
        key="pasta_traf",
    )

    st.write("")
    if st.button("Gerar Relatório de Tráfego", type="primary", use_container_width=True, key="btn_traf"):
        if not pasta_base:
            st.error("Informe a pasta de destino.")
            return
        caminho, total_pass, total_ev, total_perd = gerar_trafego_por_dia(mes, ano, pasta_base)
        if os.path.exists(caminho):
            st.success("Concluído! {} passagens · {} evasões · R$ {} perdido.".format(
                "{:,}".format(total_pass), "{:,}".format(total_ev), fmt_valor(total_perd)))
            st.success("Arquivo salvo em: {}".format(caminho))
            with open(caminho, "rb") as f:
                st.download_button(
                    "Baixar CSV", data=f.read(),
                    file_name="{}_{}_Trafego_Praca_Pedagio.csv".format(ano, str(mes).zfill(2)),
                    mime="text/csv", use_container_width=True,
                )


def pagina_sats():
    cabecalho("C-SATS — Extração AT3", "📊",
              "Consulta {} tabelas AT3 (Amostras e Ocorrências) no SQL Server.".format(len(TABELAS_SATS)))
    voltar()
    st.write("")

    mes, ano = seletor_mes_ano("sats")
    st.info("SQL Server {} · AT3_db_1 · {} tabelas · {}/{}".format(
        SQL_HOST, len(TABELAS_SATS), str(mes).zfill(2), ano))

    st.write("")
    if st.button("Gerar Relatórios SATS", type="primary", use_container_width=True, key="btn_sats"):
        resultados = gerar_sats(mes, ano)
        if resultados:
            st.session_state["sats_resultados"] = resultados
            st.session_state["sats_mes"]        = mes
            st.session_state["sats_ano"]        = ano

    if "sats_resultados" in st.session_state:
        st.divider()
        st.markdown("### Downloads SATS")
        render_downloads_sats(
            st.session_state["sats_resultados"],
            st.session_state["sats_mes"],
            st.session_state["sats_ano"],
        )

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="ARTEMIG — Cockpit",
        page_icon="🛣️",
        layout="wide",
    )
    st.markdown(CSS, unsafe_allow_html=True)

    if "pagina" not in st.session_state:
        st.session_state["pagina"] = "cockpit"

    pagina = st.session_state["pagina"]

    if pagina == "cockpit":
        pagina_cockpit()
    elif pagina == "atendimentos":
        pagina_atendimentos()
    elif pagina == "sinistros":
        pagina_sinistros()
    elif pagina == "trafego":
        pagina_trafego()
    elif pagina == "sats":
        pagina_sats()

if __name__ == "__main__":
    main()
