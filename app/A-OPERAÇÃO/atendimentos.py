import os
from datetime import datetime, date, time
from io import BytesIO

import streamlit as st
import pandas as pd
import oracledb
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Carrega variáveis de ambiente ─────────────────────────────────────────────
load_dotenv()

DB_HOST    = os.getenv("DB_HOST", "oracluster")
DB_PORT    = int(os.getenv("DB_PORT", "1521"))
DB_SERVICE = os.getenv("DB_SERVICE", "srv_bi")
DB_USER    = os.getenv("DB_USER")
DB_PASS    = os.getenv("DB_PASSWORD")
SK_EMPRESA = int(os.getenv("SK_EMPRESA", "6"))

# ── Modo THICK ────────────────────────────────────────────────────────────────
try:
    oracledb.init_oracle_client()
except Exception:
    pass

# ── Constantes ────────────────────────────────────────────────────────────────
CONCESSIONARIA = "Via Nascentes"

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}

COLUNAS_ESCOPO = [
    "CnpjConcessionaria", "NumOcorrencia", "IdVeiculo", "TipoVeiculo",
    "DataAcionamento", "KmMetrosAcionamento", "LatitudeAcionamento", "LongitudeAcionamento",
    "DataChegadaAtendimento", "DataSaidaAtendimento",
    "LatitudeAtendimento", "LongitudeAtendimento",
    "Rodovia", "KmMetrosAtendimento",
    "ViaUrbanaNaoSegregada", "Pavimentacao", "TipoServicoPrestado",
]

COR_HEADER_BG = "1F4E79"
COR_HEADER_FG = "FFFFFF"
COR_ALT_ROW   = "DCE6F1"

# ── Mapeamentos ───────────────────────────────────────────────────────────────
def map_tipo_veiculo(viatura: str) -> str:
    v = str(viatura).strip().upper()
    if v.startswith("IT-"):                             return "Inspeção de tráfego"
    if v.startswith("GL-"):                             return "Guincho leve"
    if v.startswith("GPD-"):                            return "Guincho pesado"
    if v.startswith("B-"):                              return "Ambulância Tipo C"
    if v.startswith("P-"):                              return "Caminhão pipa"
    if v.startswith("OPE-") or v.startswith("PGF-"):   return "Inspeção de tráfego"
    return ""

def map_tipo_servico(tipo_atend: str, viatura: str) -> str:
    ta = str(tipo_atend).strip().lower()
    v  = str(viatura).strip().upper()
    if ta == "cancelado":                                   return "Cancelado"
    if "nao localizado" in ta or "não localizado" in ta:   return "Não localizado"
    if v.startswith("IT-") or v.startswith("OPE-") or v.startswith("PGF-"):
        return "Inspeção (Veículo de inspeção)"
    if v.startswith("GL-"):   return "Mecânico (Guincho leve)"
    if v.startswith("GPD-"):  return "Mecânico (Guincho pesado)"
    if v.startswith("B-"):    return "Atendimento pré-hospitalar (Ambulância)"
    if v.startswith("P-"):    return "Combate a incêndio (Caminhão pipa)"
    if "inspeção" in ta or "inspecao" in ta: return "Inspeção (Veículo de inspeção)"
    if "remoção"  in ta or "remocao"  in ta: return "Mecânico (Guincho leve)"
    return ""

def safe(val) -> str:
    try:
        if pd.isna(val): return ""
    except Exception:
        pass
    return str(val).strip()

def formatar_datahora(data_val, hora_val) -> str:
    try:
        if isinstance(data_val, datetime): dt = data_val.date()
        elif isinstance(data_val, date):   dt = data_val
        else: dt = datetime.strptime(str(data_val).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return ""
    try:
        if isinstance(hora_val, time):
            hs = hora_val.strftime("%H:%M:%S")
        elif hora_val is None or (isinstance(hora_val, float) and pd.isna(hora_val)):
            hs = "00:00:00"
        else:
            hs = str(hora_val).strip()
            if len(hs) == 5: hs += ":00"
    except Exception:
        hs = "00:00:00"
    return "{}-{}".format(dt.strftime("%d/%m/%Y"), hs)

# ── Conexão Oracle ────────────────────────────────────────────────────────────
def get_connection():
    dsn = oracledb.makedsn(DB_HOST, DB_PORT, service_name=DB_SERVICE)
    return oracledb.connect(user=DB_USER, password=DB_PASS, dsn=dsn)

# ── Query ─────────────────────────────────────────────────────────────────────
QUERY = """
    SELECT
        KOCORRENCIA,
        NUMOCORRENCIA,
        DATAOCORRENCIA,
        HORAOCORRENCIA,
        COD_RODOVIA,
        RODOVIA,
        KM,
        COD_SENT,
        SENTIDO,
        COD_TIPO_ATENDIMENTO,
        DSC_TIPO_ATENDIMENTO,
        TIPO_REMOCAO,
        RECURSO,
        TEMPOCHEGADA,
        TIPOOCORRENCIA,
        DESCROCORRENCIA,
        LATITUDE,
        LONGITUDE,
        TRECHO,
        ANO,
        MES,
        DIA
    FROM STAGE_AREA.STG_ATENDIMENTOS_KCOR
    WHERE SK_EMPRESA = :sk_empresa
      AND ANO        = :ano
      AND MES        = :mes
    ORDER BY DATAOCORRENCIA, HORAOCORRENCIA
"""

def carregar_dados(mes: int, ano: int) -> pd.DataFrame:
    try:
        conn = get_connection()
        df = pd.read_sql(
            QUERY,
            conn,
            params={"sk_empresa": SK_EMPRESA, "ano": ano, "mes": mes}
        )
        conn.close()
        return df
    except Exception as e:
        st.error("Erro ao conectar ao banco de dados: {}".format(e))
        return pd.DataFrame()

# ── Transformação ─────────────────────────────────────────────────────────────
def transformar(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=COLUNAS_ESCOPO)

    rows = []
    for _, r in df.iterrows():
        viatura    = safe(r.get("RECURSO"))
        tipo_atend = safe(r.get("DSC_TIPO_ATENDIMENTO"))
        data_val   = r.get("DATAOCORRENCIA")
        hora_ac    = r.get("HORAOCORRENCIA")
        tempo_cheg = r.get("TEMPOCHEGADA")

        rows.append({
            "CnpjConcessionaria":     CONCESSIONARIA,
            "NumOcorrencia":          safe(r.get("NUMOCORRENCIA")),
            "IdVeiculo":              viatura,
            "TipoVeiculo":            map_tipo_veiculo(viatura),
            "DataAcionamento":        formatar_datahora(data_val, hora_ac),
            "KmMetrosAcionamento":    "",
            "LatitudeAcionamento":    "",
            "LongitudeAcionamento":   "",
            "DataChegadaAtendimento": formatar_datahora(data_val, tempo_cheg),
            "DataSaidaAtendimento":   "",
            "LatitudeAtendimento":    safe(r.get("LATITUDE")),
            "LongitudeAtendimento":   safe(r.get("LONGITUDE")),
            "Rodovia":                safe(r.get("RODOVIA")),
            "KmMetrosAtendimento":    safe(r.get("KM")),
            "ViaUrbanaNaoSegregada":  "",
            "Pavimentacao":           "",
            "TipoServicoPrestado":    map_tipo_servico(tipo_atend, viatura),
        })

    return pd.DataFrame(rows, columns=COLUNAS_ESCOPO)

# ── Geração do Excel ──────────────────────────────────────────────────────────
def gerar_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Atendimentos Serviços"

    hf   = Font(name="Arial", bold=True, color=COR_HEADER_FG, size=10)
    hb   = PatternFill("solid", start_color=COR_HEADER_BG)
    ha   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dfnt = Font(name="Arial", size=9)
    af   = PatternFill("solid", start_color=COR_ALT_ROW)
    ca   = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 35
    for ci, col in enumerate(COLUNAS_ESCOPO, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf; c.fill = hb; c.alignment = ha; c.border = brd

    for ri, row in enumerate(df[COLUNAS_ESCOPO].itertuples(index=False), 2):
        fill = af if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfnt; c.alignment = ca; c.border = brd
            if fill: c.fill = fill

    widths = [22,16,12,24,24,22,20,20,26,24,20,20,12,22,24,14,34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Salvar local ──────────────────────────────────────────────────────────────
def salvar_local(xlsx_bytes: bytes, mes: int, ano: int, base_dir: str) -> str:
    pasta = os.path.join(
        base_dir, "A-OPERACAO", "atendimentos_servicos",
        str(ano), "{}_{}".format(str(mes).zfill(2), MESES_PT[mes])
    )
    os.makedirs(pasta, exist_ok=True)
    nome = "{}_{}_Atendimentos_Servicos.xlsx".format(ano, str(mes).zfill(2))
    caminho = os.path.join(pasta, nome)
    with open(caminho, "wb") as f:
        f.write(xlsx_bytes)
    return caminho

# ── Interface Streamlit ───────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Relatórios Nascentes – Atendimentos",
        page_icon="🛣️",
        layout="centered",
    )

    st.title("🛣️ Relatórios Nascentes")
    st.subheader("A-OPERAÇÃO › Atendimentos e Serviços")
    st.caption("Consulta STAGE_AREA.STG_ATENDIMENTOS_KCOR e gera o escopo padronizado.")
    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        mes = st.selectbox(
            "📅 Mês do relatório",
            options=list(MESES_PT.keys()),
            format_func=lambda m: MESES_PT[m],
            index=datetime.now().month - 1,
        )
    with col2:
        ano_atual = datetime.now().year
        ano = st.number_input(
            "📅 Ano do relatório",
            min_value=2020, max_value=ano_atual + 1,
            value=ano_atual, step=1,
        )

    st.divider()
    st.markdown("### 🗄️ Fonte de dados")
    st.info(
        "Os dados serão consultados em **STAGE_AREA.STG_ATENDIMENTOS_KCOR** "
        "filtrando SK_EMPRESA = {} | Mês = {} | Ano = {}.".format(SK_EMPRESA, mes, ano),
        icon="ℹ️",
    )
    st.divider()

    st.markdown("### 📁 Pasta de destino (opcional)")
    pasta_base = st.text_input(
        "Caminho base para salvar o arquivo",
        value=os.path.join(os.path.expanduser("~"), "Relatórios Nascentes"),
    )
    st.divider()

    if st.button("⚙️  Gerar Relatório", type="primary", use_container_width=True):
        with st.spinner("Consultando banco de dados..."):
            df_raw = carregar_dados(mes, ano)

        if df_raw.empty:
            st.warning("Nenhum registro encontrado para SK_EMPRESA={}, Mês={}, Ano={}.".format(SK_EMPRESA, mes, ano))
            return

        with st.spinner("Transformando dados..."):
            df_final = transformar(df_raw)

        st.success("Consulta concluída! {} registros encontrados.".format(len(df_final)))
        st.dataframe(df_final, use_container_width=True)
        st.divider()

        xlsx_bytes = gerar_xlsx(df_final)
        st.download_button(
            label="📥 Baixar Relatório Excel",
            data=xlsx_bytes,
            file_name="{}_{}_Atendimentos_Servicos.xlsx".format(ano, str(mes).zfill(2)),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        if pasta_base:
            try:
                caminho_salvo = salvar_local(xlsx_bytes, mes, ano, pasta_base)
                st.success("Relatório salvo localmente em: {}".format(caminho_salvo))
            except Exception as e:
                st.warning("Não foi possível salvar localmente: {}".format(e))

if __name__ == "__main__":
    main()
