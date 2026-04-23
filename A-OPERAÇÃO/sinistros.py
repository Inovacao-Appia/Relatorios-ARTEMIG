import os
from datetime import datetime, date, time
from io import BytesIO

import streamlit as st
import pandas as pd
import oracledb
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Carrega variáveis de ambiente ─────────────────────────────────────────────
load_dotenv()

DB_HOST    = os.getenv("DB_HOST", "oracluster")
DB_PORT    = int(os.getenv("DB_PORT", "1521"))
DB_SERVICE = os.getenv("DB_SERVICE", "srv_bi")
DB_USER    = os.getenv("DB_USER")
DB_PASS    = os.getenv("DB_PASSWORD")
SK_EMPRESA = int(os.getenv("SK_EMPRESA", "6"))

# ── Modo THICK ────────────────────────────────────────────────────────────────
try:
    oracledb.init_oracle_client()
except Exception:
    pass

# ── Constantes ────────────────────────────────────────────────────────────────
CONCESSIONARIA = "Via Nascentes"

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}

COLUNAS_ESCOPO = [
    "CnpjConcessionaria", "NumOcorrencia", "IdEnvolvido", "DataHora",
    "Municipio", "Sre", "Rodovia", "KmMetros", "Sentido",
    "Latitude", "Longitude", "TipoPista", "TracadoVia",
    "CondicaoMetereologica", "Iluminacao", "ObstrucaoPista",
    "CausaPrincipal", "CausaSinistro", "OrdemTipoSinistro", "TipoSinistro",
    "ClassificacaoSinistro", "QuantidadeVeiculos",
    "IdVeiculo", "TipoVeiculoEnvolvidoSinistro", "MarcaModeloVeiculo",
    "AnoFabricacaoVeiculo", "PlacaVeiculo", "TipoCarga",
    "QuantidadeEnvolvidos", "TipoEnvolvido", "EstadoFisico",
    "IdadeEnvolvido", "SexoEnvolvido",
    "Ileso", "FeridoLeve", "FeridoModerado", "FeridoGrave", "Morto",
]

COR_HEADER_BG = "1F4E79"
COR_HEADER_FG = "FFFFFF"
COR_ALT_ROW   = "DCE6F1"

# ── Mapeamentos ───────────────────────────────────────────────────────────────
def _norm(txt: str) -> str:
    txt = str(txt).strip().lower()
    for a, b in [("á","a"),("à","a"),("â","a"),("ã","a"),("é","e"),("ê","e"),
                 ("í","i"),("ó","o"),("ô","o"),("õ","o"),("ú","u"),("ç","c")]:
        txt = txt.replace(a, b)
    return txt

SENTIDO_MAP = {
    "leste":  "Crescente",
    "norte":  "Crescente",
    "oeste":  "Decrescente",
    "sul":    "Decrescente",
    "l":      "Crescente",
    "n":      "Crescente",
    "o":      "Decrescente",
    "s":      "Decrescente",
}

TIPO_SINISTRO_MAP = {
    "1.0 colisao traseira":                              "Colisão traseira",
    "2.0 colisao frontal":                               "Colisão frontal",
    "3.0 colisao lateral":                               "Colisão lateral",
    "6.6 atrop. pedestre morador/trabalhador/estudante": "Atropelamento de pedestre",
    "7.0 atropelamento de animal":                       "Atropelamento de animal",
    "8.0 tombamento":                                    "Tombamento",
    "9.0 capotamento":                                   "Capotamento",
    "10.0 engavetamento":                                "Engavetamento",
    "5.2 choque com arvore":                             "Saída de pista",
    "5.3 choque com sinalizacao":                        "Saída de pista",
    "5.4 choque com defensa/barreira":                   "Saída de pista",
    "5.5 choque com elemento de drenagem":               "Saída de pista",
    "5.6 choque com talude":                             "Saída de pista",
    "5.10 objeto sobre a pista":                         "Queda de objeto",
}

TIPO_VEICULO_MAP = {
    "automovel":                    "Automóvel",
    "perua/caminhonete/utilitario": "Caminhonete",
    "moto":                         "Motocicleta",
    "caminhao":                     "Caminhão",
    "carreta":                      "Caminhão articulado",
    "onibus":                       "Ônibus",
    "van/perua":                    "Van",
    "bicicleta":                    "Bicicleta",
    "evadiu-se":                    "Não identificado",
}

CLASSIFICACAO_MAP = {
    "acidente sem vitima":       "Sem vítimas",
    "acidente com vitima":       "Com vítimas",
    "acidente com vitima fatal": "Com vítimas fatais",
}

CONDICAO_TEMPO_MAP = {
    "bom":     "Céu claro",
    "chuva":   "Chuva",
    "neblina": "Neblina/Nevoeiro",
    "nublado": "Nublado",
    "nao def": "Ignorado",
}

TRACADO_MAP = {
    "reta":            "Reta",
    "curva suave":     "Curva",
    "curva acentuada": "Curva",
    "nao def":         "",
}

CAUSA_MAP = {
    "303 perda de roda":                 "Falha mecânica",
    "402 perda de controle (impericia)": "Falta de atenção",
    "403 dormiu no volante":             "Dormiu ao volante",
    "407 levou fechada":                 "Desobediência à sinalização",
    "415 descuido do motorista":         "Falta de atenção",
    "601 objeto na pista":               "Objeto sobre a pista",
    "602 objeto na pista":               "Objeto sobre a pista",
    "204 chuva":                         "Pista escorregadia",
    "207 animal na pista":               "Animal na pista",
    "701 imprudencia do motorista":      "Velocidade incompatível",
    "nao def":                           "",
}

def map_tipo_veiculo(desc: str) -> str:
    k = _norm(desc)
    for chave, valor in TIPO_VEICULO_MAP.items():
        if chave in k:
            return valor
    return str(desc).strip()

def map_sentido(sentido_raw: str) -> str:
    return SENTIDO_MAP.get(_norm(sentido_raw), sentido_raw.strip())

def safe(val) -> str:
    try:
        if pd.isna(val): return ""
    except Exception:
        pass
    return str(val).strip()

def formatar_datahora(data_val, hora_val) -> str:
    try:
        if isinstance(data_val, datetime): dt = data_val.date()
        elif isinstance(data_val, date):   dt = data_val
        else: dt = datetime.strptime(str(data_val).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return ""
    try:
        if isinstance(hora_val, time):
            hs = hora_val.strftime("%H:%M:%S")
        elif hora_val is None or (isinstance(hora_val, float) and pd.isna(hora_val)):
            hs = "00:00:00"
        else:
            hs = str(hora_val).strip()
            if len(hs) == 5: hs += ":00"
    except Exception:
        hs = "00:00:00"
    return "{}-{}".format(dt.strftime("%d/%m/%Y"), hs)

def fmt_km(km, mt) -> str:
    try:
        return "{}+{}".format(
            str(int(float(str(km).strip()))).zfill(3),
            str(int(float(str(mt).strip()))).zfill(3)
        )
    except Exception:
        return str(km).strip()

# ── Conexão Oracle ────────────────────────────────────────────────────────────
def get_connection():
    dsn = oracledb.makedsn(DB_HOST, DB_PORT, service_name=DB_SERVICE)
    return oracledb.connect(user=DB_USER, password=DB_PASS, dsn=dsn)

# ── Queries ───────────────────────────────────────────────────────────────────
QUERY_ACIDENTES = """
    SELECT
        A.KOCORRENCIA,
        A.NUMOCORRENCIA,
        A.DATAOCORRENCIA,
        A.HORAACIDENTE,
        A.DESCROCORRENCIA,
        A.TIPOACIDENTE,
        A.CAUSAPROVAVEL,
        A.TRACADOPISTA,
        A.CONDICAOMETEOROLOGICA,
        A.LATITUDE,
        A.LONGITUDE,
        A.SIGLARODOVIA,
        A.KM,
        A.MTS,
        A.TRECHO,
        MAX(AT.SENTIDO) AS SENTIDO
    FROM STAGE_AREA.STG_KCOR_ACIDENTES2 A
    LEFT JOIN STAGE_AREA.STG_ATENDIMENTOS_KCOR AT
        ON AT.KOCORRENCIA = A.KOCORRENCIA
       AND AT.SK_EMPRESA  = A.SK_EMPRESA
    WHERE A.SK_EMPRESA = :sk_empresa
      AND EXTRACT(YEAR  FROM A.DATAOCORRENCIA) = :ano
      AND EXTRACT(MONTH FROM A.DATAOCORRENCIA) = :mes
    GROUP BY
        A.KOCORRENCIA,
        A.NUMOCORRENCIA,
        A.DATAOCORRENCIA,
        A.HORAACIDENTE,
        A.DESCROCORRENCIA,
        A.TIPOACIDENTE,
        A.CAUSAPROVAVEL,
        A.TRACADOPISTA,
        A.CONDICAOMETEOROLOGICA,
        A.LATITUDE,
        A.LONGITUDE,
        A.SIGLARODOVIA,
        A.KM,
        A.MTS,
        A.TRECHO
    ORDER BY A.DATAOCORRENCIA, A.HORAACIDENTE
"""

QUERY_VEICULOS = """
    SELECT
        V.KOCORRENCIA,
        V.COD_VEICULO_ATEND,
        V.DESC_VEICULO_ATEN,
        V.MARCAMODELO
    FROM STAGE_AREA.STG_KCOR_VEICULOS V
    WHERE V.SK_EMPRESA = :sk_empresa
      AND EXTRACT(YEAR  FROM V.DATAOCORRENCIA) = :ano
      AND EXTRACT(MONTH FROM V.DATAOCORRENCIA) = :mes
    ORDER BY V.KOCORRENCIA, V.COD_VEICULO_ATEND
"""

def carregar_dados(mes: int, ano: int):
    try:
        conn   = get_connection()
        params = {"sk_empresa": SK_EMPRESA, "ano": ano, "mes": mes}
        df_acid = pd.read_sql(QUERY_ACIDENTES, conn, params=params)
        df_veic = pd.read_sql(QUERY_VEICULOS,  conn, params=params)
        conn.close()
        return df_acid, df_veic
    except Exception as e:
        st.error("Erro ao conectar ao banco de dados: {}".format(e))
        return pd.DataFrame(), pd.DataFrame()

# ── Transformação ─────────────────────────────────────────────────────────────
def transformar(df_acid: pd.DataFrame, df_veic: pd.DataFrame) -> pd.DataFrame:
    if df_acid.empty:
        return pd.DataFrame(columns=COLUNAS_ESCOPO)

    qtd_veic = (
        df_veic.groupby("KOCORRENCIA")["COD_VEICULO_ATEND"]
        .count().reset_index()
        .rename(columns={"COD_VEICULO_ATEND": "QTD_VEIC"})
    ) if not df_veic.empty else pd.DataFrame(columns=["KOCORRENCIA", "QTD_VEIC"])

    rows = []
    for _, r in df_acid.iterrows():
        koco = r["KOCORRENCIA"]

        data_hora     = formatar_datahora(r.get("DATAOCORRENCIA"), r.get("HORAACIDENTE"))
        sentido       = map_sentido(safe(r.get("SENTIDO")))
        tracado       = TRACADO_MAP.get(_norm(safe(r.get("TRACADOPISTA"))), safe(r.get("TRACADOPISTA")))
        cond_met      = CONDICAO_TEMPO_MAP.get(_norm(safe(r.get("CONDICAOMETEOROLOGICA"))), safe(r.get("CONDICAOMETEOROLOGICA")))
        causa         = CAUSA_MAP.get(_norm(safe(r.get("CAUSAPROVAVEL"))), safe(r.get("CAUSAPROVAVEL")))
        tipo_sinistro = TIPO_SINISTRO_MAP.get(_norm(safe(r.get("TIPOACIDENTE"))), safe(r.get("TIPOACIDENTE")))
        classificacao = CLASSIFICACAO_MAP.get(_norm(safe(r.get("DESCROCORRENCIA"))), safe(r.get("DESCROCORRENCIA")))
        km_metros     = fmt_km(r.get("KM", ""), r.get("MTS", 0))
        descr_norm    = _norm(safe(r.get("DESCROCORRENCIA")))

        qtd_row = qtd_veic[qtd_veic["KOCORRENCIA"] == koco]
        qtd_v   = int(qtd_row["QTD_VEIC"].values[0]) if not qtd_row.empty else ""

        veics_oco = (
            df_veic[df_veic["KOCORRENCIA"] == koco]
            .sort_values("COD_VEICULO_ATEND")
            .reset_index(drop=True)
        ) if not df_veic.empty else pd.DataFrame()

        if veics_oco.empty:
            veics_oco = pd.DataFrame([{
                "COD_VEICULO_ATEND": 1,
                "DESC_VEICULO_ATEN": "",
                "MARCAMODELO":       "",
            }])

        for i, vei in veics_oco.iterrows():
            ileso = ferido_leve = ferido_mod = ferido_grave = morto = 0
            if i == 0:
                if "sem vitima"   in descr_norm: ileso       = 1
                elif "fatal"      in descr_norm: morto       = 1
                elif "com vitima" in descr_norm: ferido_leve = 1

            rows.append({
                "CnpjConcessionaria":           CONCESSIONARIA,
                "NumOcorrencia":                safe(r.get("NUMOCORRENCIA")),
                "IdEnvolvido":                  safe(vei.get("COD_VEICULO_ATEND")),
                "DataHora":                     data_hora,
                "Municipio":                    "",
                "Sre":                          safe(r.get("TRECHO")),
                "Rodovia":                      safe(r.get("SIGLARODOVIA")),
                "KmMetros":                     km_metros,
                "Sentido":                      sentido,
                "Latitude":                     safe(r.get("LATITUDE")),
                "Longitude":                    safe(r.get("LONGITUDE")),
                "TipoPista":                    "",
                "TracadoVia":                   tracado,
                "CondicaoMetereologica":        cond_met,
                "Iluminacao":                   "",
                "ObstrucaoPista":               "",
                "CausaPrincipal":               "Sim" if i == 0 else "Não",
                "CausaSinistro":                causa if i == 0 else "",
                "OrdemTipoSinistro":            str(i + 1),
                "TipoSinistro":                 tipo_sinistro,
                "ClassificacaoSinistro":        classificacao,
                "QuantidadeVeiculos":           qtd_v,
                "IdVeiculo":                    safe(vei.get("COD_VEICULO_ATEND")),
                "TipoVeiculoEnvolvidoSinistro": map_tipo_veiculo(safe(vei.get("DESC_VEICULO_ATEN"))),
                "MarcaModeloVeiculo":           safe(vei.get("MARCAMODELO")),
                "AnoFabricacaoVeiculo":         "",
                "PlacaVeiculo":                 "",
                "TipoCarga":                    "",
                "QuantidadeEnvolvidos":         "",
                "TipoEnvolvido":                "",
                "EstadoFisico":                 "",
                "IdadeEnvolvido":               "",
                "SexoEnvolvido":                "",
                "Ileso":                        ileso,
                "FeridoLeve":                   ferido_leve,
                "FeridoModerado":               ferido_mod,
                "FeridoGrave":                  ferido_grave,
                "Morto":                        morto,
            })

    return pd.DataFrame(rows, columns=COLUNAS_ESCOPO)

# ── Geração do Excel ──────────────────────────────────────────────────────────
def gerar_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sinistros Trânsito"

    hf   = Font(name="Arial", bold=True, color=COR_HEADER_FG, size=10)
    hb   = PatternFill("solid", start_color=COR_HEADER_BG)
    ha   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dfnt = Font(name="Arial", size=9)
    af   = PatternFill("solid", start_color=COR_ALT_ROW)
    ca   = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 35
    for ci, col in enumerate(COLUNAS_ESCOPO, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf; c.fill = hb; c.alignment = ha; c.border = brd

    for ri, row in enumerate(df[COLUNAS_ESCOPO].itertuples(index=False), 2):
        fill = af if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfnt; c.alignment = ca; c.border = brd
            if fill: c.fill = fill

    widths = [22,16,12,24,18,12,12,14,14,16,16,14,16,22,14,18,16,26,20,26,24,20,14,28,24,20,14,14,20,16,16,14,14,8,10,12,10,8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Salvar local ──────────────────────────────────────────────────────────────
def salvar_local(xlsx_bytes: bytes, mes: int, ano: int, base_dir: str) -> str:
    pasta = os.path.join(
        base_dir, "A-OPERACAO", "sinistros_transito",
        str(ano), "{}_{}".format(str(mes).zfill(2), MESES_PT[mes])
    )
    os.makedirs(pasta, exist_ok=True)
    nome = "{}_{}_Sinistros_Transito.xlsx".format(ano, str(mes).zfill(2))
    caminho = os.path.join(pasta, nome)
    with open(caminho, "wb") as f:
        f.write(xlsx_bytes)
    return caminho

# ── Interface Streamlit ───────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Relatórios Nascentes – Sinistros",
        page_icon="🛣️",
        layout="centered",
    )

    st.title("🛣️ Relatórios Nascentes")
    st.subheader("A-OPERAÇÃO › Sinistros de Trânsito")
    st.caption(
        "Consulta STAGE_AREA.STG_KCOR_ACIDENTES2 + STG_KCOR_VEICULOS + "
        "STG_ATENDIMENTOS_KCOR (sentido) e gera o escopo padronizado."
    )
    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        mes = st.selectbox(
            "📅 Mês do relatório",
            options=list(MESES_PT.keys()),
            format_func=lambda m: MESES_PT[m],
            index=datetime.now().month - 1,
        )
    with col2:
        ano_atual = datetime.now().year
        ano = st.number_input(
            "📅 Ano do relatório",
            min_value=2020, max_value=ano_atual + 1,
            value=ano_atual, step=1,
        )

    st.divider()
    st.markdown("### 🗄️ Fonte de dados")
    st.info(
        "**STAGE_AREA.STG_KCOR_ACIDENTES2** + **STG_KCOR_VEICULOS** + "
        "**STG_ATENDIMENTOS_KCOR** (sentido via KOCORRENCIA) — "
        "SK_EMPRESA = {} | Mês = {} | Ano = {}.".format(SK_EMPRESA, mes, ano),
        icon="ℹ️",
    )
    st.divider()

    st.markdown("### 📁 Pasta de destino (opcional)")
    pasta_base = st.text_input(
        "Caminho base para salvar o arquivo",
        value=os.path.join(os.path.expanduser("~"), "Relatórios Nascentes"),
    )
    st.divider()

    if st.button("⚙️  Gerar Relatório", type="primary", use_container_width=True):
        with st.spinner("Consultando banco de dados..."):
            df_acid, df_veic = carregar_dados(mes, ano)

        if df_acid.empty:
            st.warning("Nenhum acidente encontrado para SK_EMPRESA={}, Mês={}, Ano={}.".format(SK_EMPRESA, mes, ano))
            return

        with st.spinner("Cruzando tabelas e transformando dados..."):
            df_final = transformar(df_acid, df_veic)

        st.success("{} ocorrências → {} linhas geradas (uma por veículo).".format(
            df_acid.shape[0], df_final.shape[0]))
        st.dataframe(df_final, use_container_width=True)
        st.divider()

        xlsx_bytes = gerar_xlsx(df_final)
        st.download_button(
            label="📥 Baixar Relatório Excel",
            data=xlsx_bytes,
            file_name="{}_{}_Sinistros_Transito.xlsx".format(ano, str(mes).zfill(2)),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        if pasta_base:
            try:
                caminho_salvo = salvar_local(xlsx_bytes, mes, ano, pasta_base)
                st.success("Relatório salvo localmente em: {}".format(caminho_salvo))
            except Exception as e:
                st.warning("Não foi possível salvar localmente: {}".format(e))

if __name__ == "__main__":
    main()
